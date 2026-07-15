import numpy as np
from openpyxl import load_workbook

from predici_clone.control.pid import PIDController
from predici_clone.interop.excel_export import export_workbook
from predici_clone.interop.flowsheet import Flowsheet
from predici_clone.interop.reaction_graph import export_html, reaction_network
from predici_clone.kinetics.reaction import RateLaw, ReactionKind, ReactionStep
from predici_clone.reactor.energy_balance import scripted_temperature_rate
from predici_clone.reactor.feed import MassStreamTable
from predici_clone.reactor.pressure import pressure_at
from predici_clone.script.commands import ScriptCommandState, script_command_namespace
from predici_clone.script.ode_system import ODESystem


def test_pid_ode_and_full_script_commands_have_closed_loop_behavior():
    pid = PIDController(0.8, 0.2, setpoint=1.0, output_limits=(0, 2))
    measured = 0.0
    for _ in range(200):
        measured += 0.1 * (pid.update(measured, 0.1) - measured)
    system = ODESystem(("x",), {"x": "-k*x"}, {"k": 2.0})
    solved = system.solve(np.linspace(0, 1, 20), {"x": 1.0})
    state = ScriptCommandState(moments={"Mn": 4}, reactor_states={"a": {"M": 2}})
    namespace = script_command_namespace(state)

    assert abs(measured - 1) < 0.2
    assert np.isclose(solved["x"][-1], np.exp(-2), rtol=1e-5)
    assert namespace["getmn"]() == 4
    assert namespace["copyreactor"]("a", "b", 0.5) == 1


def test_feed_pressure_and_user_heat_follow_schedules():
    table = MassStreamTable(np.asarray([0, 1, 2]), np.asarray([1, 2, 3]), np.asarray([300, 320, 340]), delay=0.5)
    flow, temperature = table.value(1.0)
    builtin = scripted_temperature_rate(300, heat_capacity=2, mass=4, generated_heat=80)
    scripted = scripted_temperature_rate(300, heat_capacity=2, mass=4, generated_heat=80, script=lambda scope: scope["generated_heat"] / (scope["mass"] * scope["heat_capacity"]))

    assert (flow, temperature) == (1.5, 310)
    assert pressure_at(0.5, table=((0, 1e5), (1, 2e5))) == 1.5e5
    assert builtin == scripted == 10


def test_excel_flowsheet_and_petri_exports_are_self_contained(tmp_path):
    workbook_path = export_workbook(tmp_path / "out.xlsx", {"parity": [{"measured": 1, "predicted": 1.1}]})
    flowsheet = Flowsheet()
    flowsheet.add_unit("R1", "batch")
    flowsheet.add_unit("R2", "cstr")
    flowsheet.connect("R1", "R2")
    step = ReactionStep("grow", ReactionKind.PROPAGATION, ("M", "R"), ("R",), RateLaw("kp"))
    html = export_html(reaction_network((step,)), tmp_path / "network.html")

    assert load_workbook(workbook_path).sheetnames == ["parity"]
    assert flowsheet.export(tmp_path / "flow.json").exists()
    assert "Reaction network" in html.read_text(encoding="utf-8")
